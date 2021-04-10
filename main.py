# Librería numpy en archivos de tipo .xlsx
print("Librería numpy con Excel")

# Importaremos openpyxl para poder accerder y tabajar em archivos de tipo Excel
import openpyxl

# Importaremos de una vez también, la librería numpy con la cuál haremos diversas operaciones
import numpy as nmp

# Comenzaremos por llamar el archivo que ya cargamos llamado 'excel_1.xlsx'
wb = openpyxl.load_workbook('excel_1.xlsx')

# Ahora, para acceder a las hojas de cálculo, usamos .get_sheet_by_name('nombre de la hoja)
sheet1 = wb['Hoja1']

# Con el acceso a la hoja, podemos comenzar a sacar los datos que en ella están contenidos

# Estos son los datos correspondientes a la columna A
A2 = sheet1['A2'].value #2
A3 = sheet1['A3'].value #3
A4 = sheet1['A4'].value #1

# Estos son los datos correspondientes a la columna B
B2 = sheet1['B2'].value #0
B3 = sheet1['B3'].value #5
B4 = sheet1['B4'].value #9

print(A2,A3,A4,B2,B3,B4)
print(type(A4),type(B3))

# Procedermos a crear el archivo donde guardaremos las operaciones 

# Crear un archivo de Excel
opera_doc_1 = openpyxl.Workbook()

# Aquí asignamos una hoja de cálculo en blanco en el archivo creado en el paso anterior
hojacalculo = opera_doc_1.active

# Guardaremos lo que llevamos en el archivo con el nombre: 'opera1_excel_1.xlsx'
opera_doc_1.save('opera1_excel_1.xlsx')

# Ya creado, continuaremos diseñando algunas operaciones
hojacalculo['A1'] = ("Operaciones realizadas con la librería numpy de Pyhton")
hojacalculo['A3'] = ("Ejercicios parte 1")
hojacalculo['A4'] = 1
hojacalculo['A5'] = 2
hojacalculo['A6'] = 3
hojacalculo['A7'] = 4

# Ejercicio 1: multiplicación por pi
hojacalculo['B4'] = ("Multiplicar número de la celda 'B3' del archivo 'excel_1.xlsx' por pi:")
print("numero por pi")

# Creación da la operación
ejercicio_1_pi = B3* nmp.pi
print(ejercicio_1_pi)

# Almacenar la respuesta en una celda de 'opera1_excel_1.xlsx'
hojacalculo['C4'] = ejercicio_1_pi


# Ejercicio 2: encontrar las raíces del polinomio x**2+2x+1
hojacalculo['B5'] = ("Raíces del polinomio x**2+2x+1 tomando los coeficientes de 'excel_1.xlsx':")
print("raíces")

# Creación da la operación
ejercicio_2_raices = nmp.roots([A4,A2,A4])
print(ejercicio_2_raices)

# Almacenar la respuesta en una celda de 'opera1_excel_1.xlsx'
hojacalculo['C5'] = str(ejercicio_2_raices)


# Ejercicio 3: obtener el conjunto de números de un rango de 'A2': 2 a 'B4': 9 de 'A4': 1 en 1
hojacalculo['B6'] = ("Obtener el rango de 2 a 8, de uno en uno, tomando los números de 'excel_1.xlsx':")
print("rango")

# Creación da la operación
ejercicio_3_range = nmp.arange(A2,B4,A4)
print(ejercicio_3_range)

# Almacenar la respuesta en una celda de 'opera1_excel_1.xlsx'
hojacalculo['C6'] = str(ejercicio_3_range)


# Ejercicio 4: crear un número complejo con complex
hojacalculo['B7'] = ("Crear un número complejo tomando los números de 'excel_1.xlsx', ('A3','B2):")
print("complejo")

# Creación da la operación
ejercicio_4_complex = complex(A3,B2)
print(ejercicio_4_complex)

# Almacenar la respuesta en una celda de 'opera1_excel_1.xlsx'
hojacalculo['C7'] = str(ejercicio_4_complex)

# Guadaremos los ejercicios realizados
opera_doc_1.save('opera1_excel_1.xlsx')

print("--------"*8)

# Comenzaremos por llamar el archivo que ya cargamos llamado 'excel_2.xlsx'
wb = openpyxl.load_workbook('excel_2.xlsx')

# Ahora, para acceder a las hojas de cálculo, usamos .get_sheet_by_name('nombre de la hoja)
hoja1 = wb['Hoja1']

# Con el acceso a la hoja, podemos comenzar a sacar los datos que en ella están contenidos

# Estos son los datos correspondientes a la columna A
A2_doc2 = hoja1['A2'].value #0.14
A3_doc2 = hoja1['A3'].value #3.52
A4_doc2 = hoja1['A4'].value #1.87

# Estos son los datos correspondientes a la columna B
B2_doc2 = hoja1['B2'].value #1.09
B3_doc2 = hoja1['B3'].value #0.63
B4_doc2 = hoja1['B4'].value #0.23

# Procedermos a crear el archivo donde guardaremos las operaciones 

# Crear un archivo de Excel
opera_doc_2 = openpyxl.Workbook()

# Aquí asignamos una hoja de cálculo en blanco en el archivo creado en el paso anterior
hojacalculo2 = opera_doc_2.active

# Guardaremos lo que llevamos en el archivo con el nombre: 'opera2_excel_2.xlsx'
opera_doc_2.save('opera2_excel_2.xlsx')

# Ya creado, continuaremos diseñando algunas operaciones
hojacalculo2['A1'] = ("Operaciones realizadas con la librería numpy de Pyhton")
hojacalculo2['A3'] = ("Ejercicios parte 2")
hojacalculo2['A4'] = 1
hojacalculo2['A5'] = 2
hojacalculo2['A6'] = 3
hojacalculo2['A7'] = 4

# Ejercicio 1: calcular el coseno de un número
hojacalculo2['B4'] = ("Calcular el coseno del número de la celda 'A3' del archivo 'excel_2.xlsx':")
print("coseno")

# Creación da la operación
ejercicio_1_cos = nmp.cos(A3_doc2)
print(ejercicio_1_cos)

# Almacenar la respuesta en una celda de 'opera2_excel_2.xlsx'
hojacalculo2['C4'] = ejercicio_1_cos


# Ejercicio 2: calcular el seno de un conjunto de números
hojacalculo2['B5'] = ("Calcular el seno de un conjunto de números de las celdas 'A4','B3,'B4' del archivo 'excel_2.xlsx':")
print("seno conjunto de números")

# Creación da la operación
conjunto = [A4_doc2,B3_doc2,B4_doc2]
ejercicio_2_sen = nmp.sin(conjunto)
print(ejercicio_2_sen)

# Almacenar la respuesta en una celda de 'opera2_excel_2.xlsx'
hojacalculo2['C5'] = str(ejercicio_2_sen)


# Ejercicio 3: calcular la suma del coseno de un número + el seno de otro +3
hojacalculo2['B6'] = ("Suma del coseno del número 'B2' + el seno del número 'A2' + 3 del archivo 'excel_2.xlsx':")
print("suma coseno + seno + 3")

# Creación da la operación
ejercicio_3_suma = (nmp.cos(B2_doc2)+nmp.sin(A2_doc2) + 3) 
print(ejercicio_3_suma)

# Almacenar la respuesta en una celda de 'opera2_excel_2.xlsx'
hojacalculo2['C6'] = (ejercicio_3_suma)


# Ejercicio 4: creación de una matriz 
hojacalculo2['B7'] = ("Creación de una matriz con 'A3','B3','A4','B4' del archivo 'excel_2.xlsx':")
print("matriz")

# Creación da la operación
ejercicio_4_matriz = nmp.matrix([[A3_doc2,B3_doc2],[B4_doc2,A4_doc2]])
print(ejercicio_4_matriz)

# Almacenar la respuesta en una celda de 'opera2_excel_2.xlsx'
hojacalculo2['C7'] = str(ejercicio_4_matriz)

# Guardamos lo anterior en el archivo 'opera2_excel_2.xlsx'
opera_doc_2.save('opera2_excel_2.xlsx')